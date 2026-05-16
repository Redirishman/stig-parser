"""Tests for app.cli — argument parsing, path resolution, and end-to-end runs."""
from __future__ import annotations

import logging
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.cli import _build_parser, _resolve_paths, main

FIXTURES = Path(__file__).parent / "fixtures"


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

class TestArgumentParsing:
    def test_results_is_required(self):
        parser = _build_parser()
        with pytest.raises(SystemExit):
            parser.parse_args([])

    def test_benchmarks_is_optional(self):
        parser = _build_parser()
        # No --benchmarks flag at all — must not raise
        args = parser.parse_args(["--results", "a.xml"])
        assert args.benchmarks is None

    def test_benchmarks_accepts_empty_list(self):
        parser = _build_parser()
        args = parser.parse_args(["--results", "a.xml", "--benchmarks"])
        assert args.benchmarks == []

    def test_benchmarks_accepts_multiple_paths(self):
        parser = _build_parser()
        args = parser.parse_args(
            ["--results", "a.xml", "b.xml", "--benchmarks", "x.xml", "y.zip"]
        )
        assert args.results == ["a.xml", "b.xml"]
        assert args.benchmarks == ["x.xml", "y.zip"]

    def test_output_flag_parsed(self):
        parser = _build_parser()
        args = parser.parse_args(["--results", "a.xml", "--output", "out.xlsx"])
        assert args.output == "out.xlsx"

    def test_verbose_flag_parsed(self):
        parser = _build_parser()
        args = parser.parse_args(["--results", "a.xml", "--verbose"])
        assert args.verbose is True

    def test_verbose_default_false(self):
        parser = _build_parser()
        args = parser.parse_args(["--results", "a.xml"])
        assert args.verbose is False


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------

class TestResolvePaths:
    def test_directory_resolves_to_xml_files(self, tmp_path):
        (tmp_path / "a.xml").write_text("<a/>")
        (tmp_path / "b.xml").write_text("<b/>")
        (tmp_path / "skip.txt").write_text("not xml")
        paths = _resolve_paths([str(tmp_path)])
        names = sorted(p.name for p in paths)
        assert names == ["a.xml", "b.xml"]

    def test_directory_with_zip_extension_filter(self, tmp_path):
        (tmp_path / "a.xml").write_text("<a/>")
        (tmp_path / "b.zip").write_bytes(b"PK\x03\x04")
        paths = _resolve_paths([str(tmp_path)], extensions=(".xml", ".zip"))
        names = sorted(p.name for p in paths)
        assert names == ["a.xml", "b.zip"]

    def test_explicit_file_passes_through(self, tmp_path):
        f = tmp_path / "single.xml"
        f.write_text("<x/>")
        paths = _resolve_paths([str(f)])
        assert paths == [f]

    def test_glob_pattern_expanded(self, tmp_path):
        (tmp_path / "scan1.xml").write_text("<x/>")
        (tmp_path / "scan2.xml").write_text("<x/>")
        (tmp_path / "other.xml").write_text("<x/>")
        paths = _resolve_paths([str(tmp_path / "scan*.xml")])
        names = sorted(p.name for p in paths)
        assert names == ["scan1.xml", "scan2.xml"]


# ---------------------------------------------------------------------------
# End-to-end main() invocations
# ---------------------------------------------------------------------------

class TestMainSeparateBenchmarks:
    """Traditional flow: --results + --benchmarks both supplied."""

    def test_separate_benchmark_produces_workbook(self, tmp_path):
        out = tmp_path / "out.xlsx"
        rc = main([
            "--results", str(FIXTURES / "scc_results.xml"),
            "--benchmarks", str(FIXTURES / "sample_benchmark.xml"),
            "--output", str(out),
        ])
        assert rc == 0
        assert out.exists()
        wb = load_workbook(str(out))
        assert {"Findings", "Summary"} <= set(wb.sheetnames)
        # Findings sheet has data rows beyond the header
        assert wb["Findings"].max_row >= 2


class TestMainOptionalBenchmarks:
    """SCC self-contained flow: --benchmarks omitted, results used for both sides."""

    def test_no_benchmarks_flag_uses_results_files(self, tmp_path, caplog):
        out = tmp_path / "out.xlsx"
        with caplog.at_level(logging.INFO, logger="app.cli"):
            rc = main([
                "--results", str(FIXTURES / "scc_results.xml"),
                "--output", str(out),
            ])
        assert rc == 0
        assert out.exists()
        # The fallback message should appear in the log
        assert any(
            "No --benchmarks supplied" in r.message for r in caplog.records
        ), "expected fallback INFO message when --benchmarks omitted"

    def test_empty_benchmarks_flag_also_uses_results_files(self, tmp_path):
        """`--benchmarks` with no values should behave like omitting the flag."""
        out = tmp_path / "out.xlsx"
        rc = main([
            "--results", str(FIXTURES / "scc_results.xml"),
            "--benchmarks",
            "--output", str(out),
        ])
        assert rc == 0
        assert out.exists()


class TestMainErrorCases:
    def test_empty_results_dir_exits_nonzero(self, tmp_path, caplog):
        """An empty directory has no .xml files → 'No results files found' → exit 1."""
        empty = tmp_path / "empty"
        empty.mkdir()
        out = tmp_path / "out.xlsx"
        with caplog.at_level(logging.ERROR, logger="app.cli"):
            rc = main([
                "--results", str(empty),
                "--benchmarks", str(FIXTURES / "sample_benchmark.xml"),
                "--output", str(out),
            ])
        assert rc == 1
        assert not out.exists()
        assert any(
            "No results files found" in r.message for r in caplog.records
        )
