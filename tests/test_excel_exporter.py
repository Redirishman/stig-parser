"""Tests for ExcelExporter."""
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.exporters.excel_exporter import (
    ExcelExporter,
    _FORMULA_PREFIXES,
    _formula_quote,
    _sanitize_cell,
)
from app.parsers.base import Finding


def _finding(
    status: str = "Open",
    severity: str = "CAT I",
    server: str = "SERVER01",
    ip: str = "10.0.0.1",
    stig_title: str = "Windows Server 2022 STIG",
    vuln_id: str = "V-254239",
    rule_id: str = "SV-254239r945408_rule",
    check: str = "Check text here.",
    fix: str = "Fix text here.",
) -> Finding:
    return Finding(
        stig_title=stig_title,
        vuln_id=vuln_id,
        rule_id=rule_id,
        severity=severity,
        status=status,
        server=server,
        ip_address=ip,
        check_text=check,
        fix_text=fix,
    )


@pytest.fixture()
def sample_findings():
    return [
        _finding("Open",        "CAT I",   "SERVER01"),
        _finding("Not Reviewed","CAT II",  "SERVER02"),
        _finding("Error",       "CAT III", "SERVER01"),
        _finding("Unknown",     "CAT II",  "SERVER02"),
    ]


@pytest.fixture()
def workbook(tmp_path, sample_findings):
    exporter = ExcelExporter()
    path = tmp_path / "findings.xlsx"
    exporter.export(sample_findings, path)
    return load_workbook(str(path))


class TestFindingsSheet:
    def test_sheet_exists(self, workbook):
        assert "Findings" in workbook.sheetnames

    def test_header_row(self, workbook):
        ws = workbook["Findings"]
        headers = [ws.cell(1, c).value for c in range(1, 10)]
        assert "STIG Title" in headers
        assert "Severity" in headers
        assert "Status" in headers
        assert "Check Text" in headers
        assert "Fix Text" in headers

    def test_data_rows(self, workbook):
        ws = workbook["Findings"]
        # 4 data rows + 1 header = max_row 5
        assert ws.max_row == 5

    def test_freeze_pane(self, workbook):
        ws = workbook["Findings"]
        assert ws.freeze_panes == "A2"

    def test_auto_filter_set(self, workbook):
        ws = workbook["Findings"]
        assert ws.auto_filter.ref is not None


class TestSummarySheet:
    def test_sheet_exists(self, workbook):
        assert "Summary" in workbook.sheetnames

    def test_table1_header(self, workbook):
        ws = workbook["Summary"]
        # "Findings by Severity" should appear somewhere in col 1
        col1_values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert "Findings by Severity" in col1_values

    def test_table2_header(self, workbook):
        ws = workbook["Summary"]
        col1_values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert "Findings by Server" in col1_values

    def test_table3_header(self, workbook):
        ws = workbook["Summary"]
        col1_values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        assert "Findings by STIG" in col1_values

    def test_countifs_formulas_present(self, workbook):
        ws = workbook["Summary"]
        formula_cells = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=COUNTIFS"):
                    formula_cells.append(cell.value)
        assert len(formula_cells) > 0, "No COUNTIFS formulas found in Summary sheet"

    def test_countifs_reference_findings_sheet(self, workbook):
        ws = workbook["Summary"]
        formula_cells = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=COUNTIFS")
        ]
        assert len(formula_cells) > 0
        assert all("Findings!" in v for v in formula_cells)


class TestFormulaInjection:
    """Guards the CWE-1236 fix: scan-derived text must never execute as a
    formula in the accreditation-facing workbook."""

    # ── _formula_quote: values interpolated into COUNTIFS criteria ────────
    def test_formula_quote_wraps_plain_value(self):
        assert _formula_quote("SERVER01") == '"SERVER01"'

    def test_formula_quote_escapes_embedded_double_quote(self):
        # A bare " would close the criteria literal and let the rest of the
        # value become live formula text; it must be doubled per Excel rules.
        assert _formula_quote('a"b') == '"a""b"'

    def test_formula_quote_neutralizes_criteria_breakout(self):
        # Classic breakout payload: close the string, inject a call, reopen.
        payload = '","")+cmd|calc!A1&COUNTIF(A:A,"'
        quoted = _formula_quote(payload)
        # Every input quote is doubled; no lone " survives to break the literal.
        assert quoted.count('""') == payload.count('"')
        assert quoted.startswith('"') and quoted.endswith('"')

    def test_formula_quote_stringifies_non_str(self):
        assert _formula_quote(42) == '"42"'

    # ── _sanitize_cell: leading formula-trigger characters ────────────────
    @pytest.mark.parametrize("prefix", _FORMULA_PREFIXES)
    def test_sanitize_prefixes_dangerous_leading_char(self, prefix):
        payload = f"{prefix}HYPERLINK(\"http://evil\")"
        assert _sanitize_cell(payload) == "'" + payload

    def test_sanitize_leaves_safe_value_untouched(self):
        assert _sanitize_cell("SERVER01") == "SERVER01"

    def test_sanitize_passes_non_str_through(self):
        assert _sanitize_cell(7) == 7

    # ── End-to-end: the saved workbook is not injectable ──────────────────
    def test_malicious_server_is_escaped_in_summary_countifs(self, tmp_path):
        evil = 'HOST","")+SUM(1,1)+COUNTIF(A:A,"'
        exporter = ExcelExporter()
        path = tmp_path / "evil.xlsx"
        exporter.export([_finding(server=evil)], path)
        wb = load_workbook(str(path))
        ws = wb["Summary"]

        countifs = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=COUNTIFS")
        ]
        server_formulas = [f for f in countifs if "$F:$F" in f]
        assert server_formulas, "no By-Server COUNTIFS formula found"
        for f in server_formulas:
            # The payload's quotes must be doubled inside the formula; a lone
            # HOST" fragment would mean the literal was broken out of.
            assert 'HOST""' in f
            assert 'HOST","' not in f

    def test_malicious_leading_char_is_neutralized_in_findings_sheet(self, tmp_path):
        evil = '=1+1'
        exporter = ExcelExporter()
        path = tmp_path / "lead.xlsx"
        exporter.export([_finding(server=evil)], path)
        wb = load_workbook(str(path))
        ws = wb["Findings"]

        # Server is column F (6); data starts at row 2.
        cell = ws.cell(row=2, column=6).value
        assert cell == "'=1+1", "leading '=' was not neutralized to literal text"
        # data_type 's' (string), not 'f' (formula) — Excel won't evaluate it.
        assert ws.cell(row=2, column=6).data_type == "s"


class TestErrorCases:
    def test_empty_findings_raises(self, tmp_path):
        exporter = ExcelExporter()
        with pytest.raises(ValueError, match="No findings"):
            exporter.export([], tmp_path / "empty.xlsx")

    def test_output_file_created(self, tmp_path):
        exporter = ExcelExporter()
        path = tmp_path / "output.xlsx"
        exporter.export([_finding()], path)
        assert path.exists()
