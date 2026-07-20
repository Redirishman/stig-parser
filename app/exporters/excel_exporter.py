"""Excel workbook generation using openpyxl."""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.parsers.base import Finding

log = logging.getLogger(__name__)

# Excel/Calc treat a leading =, +, -, @, or control char as a formula.
# Scan-derived text (hostname, check/fix text) is attacker-controllable, so
# any such value is prefixed with an apostrophe to force literal-text display.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "|", "\t", "\r")


def _sanitize_cell(value: object) -> object:
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _formula_quote(value: object) -> str:
    """Return an Excel double-quoted string literal with embedded quotes escaped.

    Values interpolated into COUNTIFS criteria (hostname, STIG title) are
    scan-derived and attacker-controllable. Excel escapes a literal " inside a
    quoted string as "". Without this, a value containing a " breaks out of the
    criteria literal and lets the upload author inject arbitrary formula text
    into the accreditation-facing Summary sheet (CWE-1236).
    """
    return '"' + str(value).replace('"', '""') + '"'


_FILL_CAT_I = PatternFill("solid", fgColor="FFCCCC")
_FILL_CAT_II = PatternFill("solid", fgColor="FFEB9C")
_FILL_CAT_III = PatternFill("solid", fgColor="C6EFCE")
_SEVERITY_FILL = {"CAT I": _FILL_CAT_I, "CAT II": _FILL_CAT_II, "CAT III": _FILL_CAT_III}

_HEADER_FONT = Font(name="Arial", size=10, bold=True)
_BODY_FONT = Font(name="Arial", size=10)

# (header_label, Finding_attr, max_col_width)
_FINDINGS_COLS: list[tuple[str, str, int]] = [
    ("STIG Title",  "stig_title",  50),
    ("Vuln ID",     "vuln_id",     12),
    ("Rule ID",     "rule_id",     40),
    ("Severity",    "severity",    10),
    ("Status",      "status",      14),
    ("Server",      "server",      30),
    ("IP Address",  "ip_address",  18),
    ("Check Text",  "check_text",  80),
    ("Fix Text",    "fix_text",    80),
]

_WRAP_HEADERS = {"Check Text", "Fix Text"}

# Findings sheet column letters (A=1 … I=9)
_COL_STIG     = "A"   # col 1
_COL_SEVERITY = "D"   # col 4
_COL_STATUS   = "E"   # col 5
_COL_SERVER   = "F"   # col 6
_COL_IP       = "G"   # col 7


class ExcelExporter:
    """Generate an Excel workbook from a list of Finding objects."""

    def export(self, findings: list[Finding], output_path: Path) -> Path:
        """Write *findings* to *output_path* and return it.

        Raises ValueError when *findings* is empty.
        """
        if not findings:
            raise ValueError("No findings to export — workbook not generated.")

        wb = Workbook()
        findings_ws = wb.active
        findings_ws.title = "Findings"
        self._write_findings(findings_ws, findings)

        summary_ws = wb.create_sheet("Summary")
        self._write_summary(summary_ws, findings)

        wb.save(str(output_path))
        log.info("Workbook written to %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Findings sheet
    # ------------------------------------------------------------------

    def _write_findings(self, ws, findings: list[Finding]) -> None:
        # Header row
        for col_idx, (header, _, _mw) in enumerate(_FINDINGS_COLS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_FINDINGS_COLS))}1"

        # Data rows
        for row_idx, finding in enumerate(findings, start=2):
            for col_idx, (header, attr, _mw) in enumerate(_FINDINGS_COLS, start=1):
                value = getattr(finding, attr, "")
                wrap = header in _WRAP_HEADERS
                cell = ws.cell(row=row_idx, column=col_idx, value=_sanitize_cell(value))
                cell.font = _BODY_FONT
                cell.alignment = Alignment(
                    wrap_text=wrap,
                    vertical="top" if wrap else "center",
                )
                if header == "Severity":
                    fill = _SEVERITY_FILL.get(value)
                    if fill:
                        cell.fill = fill

        # Column widths — measure actual content, cap at max_width
        for col_idx, (header, attr, max_w) in enumerate(_FINDINGS_COLS, start=1):
            col_letter = get_column_letter(col_idx)
            measured = len(header)
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value or ""
                measured = max(measured, min(len(str(val).split("\n")[0]), max_w))
            ws.column_dimensions[col_letter].width = min(measured + 2, max_w)

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------

    def _write_summary(self, ws, findings: list[Finding]) -> None:
        f = "Findings"  # sheet reference prefix
        severities = ["CAT I", "CAT II", "CAT III"]
        statuses   = ["Open", "Not Reviewed", "Error", "Unknown"]

        def countifs2(col_a: str, crit_a: str, col_b: str, crit_b: str) -> str:
            return (
                f'=COUNTIFS({f}!${col_a}:${col_a},{crit_a},'
                f'{f}!${col_b}:${col_b},{crit_b})'
            )

        def h(ws, row, col, text):
            c = ws.cell(row=row, column=col, value=text)
            c.font = _HEADER_FONT
            return c

        def b(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font = _BODY_FONT
            return c

        row = 1

        # ── Table 1: By Severity ──────────────────────────────────────
        h(ws, row, 1, "Findings by Severity")
        row += 1
        for ci, lbl in enumerate(["Severity", *statuses, "Total"], 1):
            h(ws, row, ci, lbl)
        row += 1

        for sev in severities:
            b(ws, row, 1, sev)
            for ci, stat in enumerate(statuses, 2):
                b(ws, row, ci, countifs2(_COL_SEVERITY, f'"{sev}"', _COL_STATUS, f'"{stat}"'))
            b(ws, row, 6, f"=SUM(B{row}:E{row})")
            row += 1
        row += 1  # spacer

        # ── Table 2: By Server ────────────────────────────────────────
        h(ws, row, 1, "Findings by Server")
        row += 1
        for ci, lbl in enumerate(["Server", "IP Address", *severities, "Total"], 1):
            h(ws, row, ci, lbl)
        row += 1

        for server, ip in _unique_pairs(findings, "server", "ip_address"):
            b(ws, row, 1, _sanitize_cell(server))
            b(ws, row, 2, _sanitize_cell(ip))
            for ci, sev in enumerate(severities, 3):
                b(ws, row, ci, countifs2(_COL_SERVER, _formula_quote(server), _COL_SEVERITY, f'"{sev}"'))
            b(ws, row, 6, f"=SUM(C{row}:E{row})")
            row += 1
        row += 1  # spacer

        # ── Table 3: By STIG ──────────────────────────────────────────
        h(ws, row, 1, "Findings by STIG")
        row += 1
        for ci, lbl in enumerate(["STIG Title", *severities, "Total"], 1):
            h(ws, row, ci, lbl)
        row += 1

        for stig in _unique_values(findings, "stig_title"):
            b(ws, row, 1, _sanitize_cell(stig))
            for ci, sev in enumerate(severities, 2):
                b(ws, row, ci, countifs2(_COL_STIG, _formula_quote(stig), _COL_SEVERITY, f'"{sev}"'))
            b(ws, row, 5, f"=SUM(B{row}:D{row})")
            row += 1
        row += 1  # spacer

        # ── Footer note ───────────────────────────────────────────────
        note_cell = ws.cell(
            row=row,
            column=1,
            value=(
                "Note: Counts use COUNTIFS and reflect all data. "
                "Filtering the Findings sheet does not update these counts. "
                "See README for details."
            ),
        )
        note_cell.font = Font(name="Arial", size=9, italic=True, color="808080")
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=6,
        )

        # Auto-width summary columns
        for col_idx in range(1, 7):
            col_letter = get_column_letter(col_idx)
            max_len = 10
            for r in range(1, row + 1):
                val = ws.cell(row=r, column=col_idx).value or ""
                if not str(val).startswith("="):
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _unique_pairs(
    findings: list[Finding], attr1: str, attr2: str
) -> list[tuple[str, str]]:
    seen: dict[tuple[str, str], None] = {}
    for f in findings:
        key = (getattr(f, attr1, ""), getattr(f, attr2, ""))
        seen.setdefault(key, None)
    return list(seen)


def _unique_values(findings: list[Finding], attr: str) -> list[str]:
    seen: dict[str, None] = {}
    for f in findings:
        seen.setdefault(getattr(f, attr, ""), None)
    return [v for v in seen if v]
